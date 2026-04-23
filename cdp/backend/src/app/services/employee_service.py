"""员工业务逻辑服务"""

import io
from datetime import datetime
from typing import Optional

from fastapi import HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.models.employee import Employee, EmployeeStatus
from app.models.employee_transfer import EmployeeTransfer
from app.repositories.employee_repo import EmployeeRepository
from app.repositories.org_unit_repo import OrgUnitRepository
from app.schemas.employee import (
    EmployeeCreate,
    EmployeeListResponse,
    EmployeeResponse,
    EmployeeTransferCreate,
    EmployeeUpdate,
)


class EmployeeService:
    """员工服务"""

    def __init__(self, db: Session, permitted_unit_ids: Optional[list[int]] = None):
        self.db = db
        self.repo = EmployeeRepository(db, permitted_unit_ids)
        self.unit_repo = OrgUnitRepository(db)

    # ------------------------------------------------------------------ #
    # 查询
    # ------------------------------------------------------------------ #

    def get_employee(self, employee_id: int) -> Employee:
        emp = self.repo.find_by_id(employee_id)
        if not emp:
            raise HTTPException(status_code=404, detail="员工不存在")
        return emp

    def list_employees(
        self,
        unit_id: Optional[int] = None,
        include_subordinates: bool = False,
        status: Optional[EmployeeStatus] = None,
        keyword: Optional[str] = None,
        page: int = 1,
        page_size: int = 20,
    ) -> EmployeeListResponse:
        unit_ids: Optional[list[int]] = None
        if unit_id is not None:
            if include_subordinates:
                unit_ids = self.unit_repo.get_descendant_ids(unit_id)
            else:
                unit_ids = [unit_id]

        items, total = self.repo.find_by_unit_ids(
            unit_ids=unit_ids or [],
            status=status,
            keyword=keyword,
            page=page,
            page_size=page_size,
        )

        response_items = []
        for emp in items:
            secondary_unit_ids = [s.unit_id for s in emp.secondary_units]
            emp_response = EmployeeResponse(
                id=emp.id,
                employee_no=emp.employee_no,
                name=emp.name,
                phone=emp.phone,
                email=emp.email,
                position=emp.position,
                primary_unit_id=emp.primary_unit_id,
                status=emp.status,
                entry_date=emp.entry_date,
                dimission_date=emp.dimission_date,
                secondary_unit_ids=secondary_unit_ids,
                user_mapping=emp.user_mapping,
                created_at=emp.created_at,
                updated_at=emp.updated_at,
            )
            response_items.append(emp_response)

        return EmployeeListResponse(
            items=response_items,
            total=total,
            page=page,
            page_size=page_size,
        )

    # ------------------------------------------------------------------ #
    # 创建 / 更新 / 删除
    # ------------------------------------------------------------------ #

    def create_employee(self, data: EmployeeCreate) -> Employee:
        # 工号唯一检查
        if self.repo.find_by_employee_no(data.employee_no):
            raise HTTPException(status_code=400, detail=f"工号 {data.employee_no!r} 已存在")

        # user 绑定检查
        if data.user_id is not None:
            existing_mapping = self.repo.find_mapping_by_user_id(data.user_id)
            if existing_mapping:
                raise HTTPException(status_code=400, detail="该账号已绑定其他员工")

        employee = Employee(
            employee_no=data.employee_no,
            name=data.name,
            phone=data.phone,
            email=data.email,
            position=data.position,
            primary_unit_id=data.primary_unit_id,
            status=EmployeeStatus.onboarding,
            entry_date=data.entry_date,
        )
        self.repo.create(employee)

        # 辅属部门
        if data.secondary_unit_ids:
            self.repo.set_secondary_units(employee, data.secondary_unit_ids)

        # 绑定账号
        if data.user_id is not None:
            self.repo.bind_user(employee.id, data.user_id)

        self.repo.commit()
        return self.repo.find_by_id(employee.id)  # 重新加载关联

    def update_employee(self, employee_id: int, data: EmployeeUpdate) -> Employee:
        employee = self.get_employee(employee_id)

        if data.name is not None:
            employee.name = data.name
        if data.phone is not None:
            employee.phone = data.phone
        if data.email is not None:
            employee.email = data.email
        if data.position is not None:
            employee.position = data.position
        if data.primary_unit_id is not None:
            employee.primary_unit_id = data.primary_unit_id
        if data.entry_date is not None:
            employee.entry_date = data.entry_date

        self.repo.update(employee)

        if data.secondary_unit_ids is not None:
            self.repo.set_secondary_units(employee, data.secondary_unit_ids)

        self.repo.commit()
        return self.repo.find_by_id(employee_id)

    def delete_employee(self, employee_id: int) -> None:
        employee = self.get_employee(employee_id)
        self.repo.soft_delete(employee)
        self.repo.commit()

    # ------------------------------------------------------------------ #
    # 账号绑定 / 解绑
    # ------------------------------------------------------------------ #

    def bind_user(self, employee_id: int, user_id: int) -> Employee:
        # 检查员工是否存在
        self.get_employee(employee_id)

        # 检查员工是否已绑定账号
        if self.repo.find_mapping_by_employee_id(employee_id):
            raise HTTPException(status_code=400, detail="该员工已绑定账号，请先解绑")

        # 检查账号是否已被绑定
        if self.repo.find_mapping_by_user_id(user_id):
            raise HTTPException(status_code=400, detail="该账号已绑定其他员工")

        self.repo.bind_user(employee_id, user_id)
        self.repo.commit()
        return self.repo.find_by_id(employee_id)

    def unbind_user(self, employee_id: int) -> Employee:
        employee = self.get_employee(employee_id)
        self.repo.unbind_user(employee)
        self.repo.commit()
        return self.repo.find_by_id(employee_id)

    # ------------------------------------------------------------------ #
    # 调动 (Task 5.4)
    # ------------------------------------------------------------------ #

    def initiate_transfer(self, employee_id: int, data: EmployeeTransferCreate) -> EmployeeTransfer:
        employee = self.get_employee(employee_id)

        # 校验目标部门存在
        to_unit = self.unit_repo.find_by_id(data.to_unit_id)
        if not to_unit:
            raise HTTPException(status_code=404, detail="目标部门不存在")

        from_unit_id = employee.primary_unit_id

        transfer = EmployeeTransfer(
            employee_id=employee_id,
            from_unit_id=from_unit_id,
            to_unit_id=data.to_unit_id,
            transfer_type=data.transfer_type,
            effective_date=data.effective_date,
            reason=data.reason,
        )
        transfer = self.repo.create_transfer(transfer)

        # 更新员工状态为调动中
        employee.status = EmployeeStatus.transferring
        self.repo.update(employee)

        # OA 审批 stub（跨部门调动触发外部审批，此处仅记录日志）
        if from_unit_id != data.to_unit_id:
            self._trigger_oa_approval_stub(transfer)

        self.repo.commit()
        return transfer

    def approve_transfer(self, employee_id: int, transfer_id: int) -> Employee:
        """执行调动：更新员工主部门并置为在职"""
        employee = self.get_employee(employee_id)

        transfers = self.repo.find_transfers_by_employee(employee_id)
        transfer = next((t for t in transfers if t.id == transfer_id), None)
        if not transfer:
            raise HTTPException(status_code=404, detail="调动记录不存在")

        employee.primary_unit_id = transfer.to_unit_id
        employee.status = EmployeeStatus.on_job
        self.repo.update(employee)
        self.repo.commit()
        return self.repo.find_by_id(employee_id)

    def get_transfer_history(self, employee_id: int) -> list[EmployeeTransfer]:
        self.get_employee(employee_id)  # 确保存在
        return self.repo.find_transfers_by_employee(employee_id)

    def confirm_onboarding(self, employee_id: int) -> Employee:
        """确认入职，将员工状态从 onboarding 改为 on_job"""
        employee = self.get_employee(employee_id)
        if employee.status != EmployeeStatus.onboarding:
            raise HTTPException(status_code=400, detail="只有入职中的员工可以确认入职")
        employee.status = EmployeeStatus.on_job
        self.repo.update(employee)
        self.repo.commit()
        return self.repo.find_by_id(employee_id)

    def _trigger_oa_approval_stub(self, transfer: EmployeeTransfer) -> None:
        """OA 审批触发 stub（跨部门调动）"""
        import logging

        logger = logging.getLogger(__name__)
        logger.info(
            "OA审批 stub: 员工 %s 调动 from %s to %s (type=%s, effective=%s)",
            transfer.employee_id,
            transfer.from_unit_id,
            transfer.to_unit_id,
            transfer.transfer_type,
            transfer.effective_date,
        )

    # ------------------------------------------------------------------ #
    # Excel 批量导入 (Task 5.5)
    # ------------------------------------------------------------------ #

    async def import_employees_from_excel(self, file: UploadFile) -> dict:
        """从 Excel 文件批量导入员工"""
        try:
            import openpyxl
        except ImportError:
            raise HTTPException(status_code=500, detail="缺少依赖 openpyxl")

        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active

        headers = [cell.value for cell in ws[1]]
        required_headers = {"工号", "姓名"}
        if not required_headers.issubset(set(headers)):
            raise HTTPException(status_code=400, detail=f"Excel 缺少必要列: {required_headers}")

        col_map = {v: i for i, v in enumerate(headers)}
        success_count = 0
        error_rows: list[dict] = []

        # 预加载所有已存在的工号，避免 N+1 查询
        existing_nos = self.repo.find_all_employee_nos()

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                employee_no = str(row[col_map["工号"]] or "").strip()
                name = str(row[col_map["姓名"]] or "").strip()
                if not employee_no or not name:
                    error_rows.append({"row": row_idx, "error": "工号或姓名为空"})
                    continue

                if employee_no in existing_nos:
                    error_rows.append({"row": row_idx, "error": f"工号 {employee_no!r} 已存在"})
                    continue

                phone = str(row[col_map.get("电话", -1)] or "").strip() if "电话" in col_map else None
                email = str(row[col_map.get("邮箱", -1)] or "").strip() if "邮箱" in col_map else None
                position = str(row[col_map.get("职位", -1)] or "").strip() if "职位" in col_map else None

                employee = Employee(
                    employee_no=employee_no,
                    name=name,
                    phone=phone or None,
                    email=email or None,
                    position=position or None,
                    status=EmployeeStatus.onboarding,
                )
                self.repo.create(employee)
                existing_nos.add(employee_no)  # 防止同一批次内重复工号
                success_count += 1
            except Exception as e:
                error_rows.append({"row": row_idx, "error": str(e)})

        self.repo.commit()
        return {"success": success_count, "errors": error_rows}

    # ------------------------------------------------------------------ #
    # Excel 导出 (Task 5.6)
    # ------------------------------------------------------------------ #

    def export_employees_to_excel(
        self,
        unit_id: Optional[int] = None,
        include_subordinates: bool = False,
        status: Optional[EmployeeStatus] = None,
    ) -> StreamingResponse:
        """导出员工列表为 Excel，使用 StreamingResponse"""
        try:
            import openpyxl
            from openpyxl.styles import Font
        except ImportError:
            raise HTTPException(status_code=500, detail="缺少依赖 openpyxl")

        unit_ids: Optional[list[int]] = None
        if unit_id is not None:
            if include_subordinates:
                unit_ids = self.unit_repo.get_descendant_ids(unit_id)
            else:
                unit_ids = [unit_id]

        employees = self.repo.find_all_for_export(unit_ids=unit_ids, status=status)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "员工列表"

        headers = ["工号", "姓名", "电话", "邮箱", "职位", "状态", "入职日期", "离职日期"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)

        for emp in employees:
            ws.append(
                [
                    emp.employee_no,
                    emp.name,
                    emp.phone or "",
                    emp.email or "",
                    emp.position or "",
                    emp.status.value,
                    str(emp.entry_date) if emp.entry_date else "",
                    str(emp.dimission_date) if emp.dimission_date else "",
                ]
            )

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"employees_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
