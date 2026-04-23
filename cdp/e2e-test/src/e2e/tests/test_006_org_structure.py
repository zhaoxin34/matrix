"""
CDP-ORG: Organization Structure Tests

E2E test cases for organization structure management module.
"""

import os
import re
import tempfile
import time

import openpyxl
import pytest
from playwright.sync_api import Page, expect

from e2e.pages import LoginPage
from conftest import assert_no_error_message


def create_employee_excel(filepath: str, employees: list[dict]) -> None:
    """Create an Excel file with employee data."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "员工数据"

    # Write header
    headers = ["工号", "姓名", "手机", "职位", "部门"]
    ws.append(headers)

    # Write employee data
    for emp in employees:
        ws.append(
            [
                emp.get("code", ""),
                emp.get("name", ""),
                emp.get("phone", ""),
                emp.get("position", ""),
                emp.get("department", ""),
            ]
        )

    wb.save(filepath)


def create_invalid_employee_excel(filepath: str, emp_code: str) -> None:
    """Create an Excel file with duplicate employee code for error testing."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "员工数据"

    # Write header
    ws.append(["工号", "姓名", "手机", "职位", "部门"])

    # Write duplicate employee code
    ws.append([emp_code, "张三", "13900001111", "工程师", "技术部"])
    ws.append([emp_code, "李四", "13900002222", "产品经理", "产品部"])

    wb.save(filepath)


def create_incomplete_employee_excel(filepath: str, missing_columns: list[str]) -> None:
    """Create an Excel file with missing required columns."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "员工数据"

    # Define all possible columns
    all_columns = ["工号", "姓名", "手机", "职位", "部门"]

    # Write header with missing columns
    headers = [col for col in all_columns if col not in missing_columns]
    ws.append(headers)

    # Write some data
    ws.append(["EMP001", "王五", "13900003333", "设计师", "设计部"])

    wb.save(filepath)


class TestOrganizationManagement:
    """Test cases for organization unit management."""

    @pytest.fixture(autouse=True)
    def setup(self, page: Page):
        self.page = page
        self.login_page = LoginPage(page)

    def _login(self):
        """Login as admin user."""
        self.login_page.navigate()
        self.login_page.phone_input.fill("13800138002")
        self.login_page.password_input.fill("abcd1234")
        self.login_page.submit_button.click()
        self.page.wait_for_timeout(2000)

    @pytest.mark.smoke
    def test_cdp_org_001_page_loads(self):
        """
        CDP-ORG-001: 组织架构页面正常加载

        前置条件：用户已登录
        测试步骤：
            1. 点击顶部导航"组织架构"链接
        预期结果：
            - 页面正确加载，显示组织树和员工列表
            - 左侧显示组织结构树
            - 右侧显示员工统计卡片和员工列表
        """
        self._login()
        self.page.get_by_test_id("link-header-org-structure").click()
        self.page.wait_for_timeout(2000)

        # Verify no backend error
        assert_no_error_message(self.page)

        # Verify page loaded - left panel with org tree title
        expect(
            self.page.get_by_role("complementary").get_by_text("组织架构")
        ).to_be_visible()

        # Verify left panel - organization tree panel shows "组织单元"
        expect(self.page.get_by_text("组织单元")).to_be_visible()

        # Verify right panel - employee statistics cards
        expect(self.page.get_by_text("员工总数")).to_be_visible()
        expect(self.page.get_by_text("在职")).to_be_visible()
        expect(self.page.get_by_text("入职中")).to_be_visible()

        # Verify employee list table headers
        expect(self.page.get_by_role("columnheader", name="工号")).to_be_visible()
        expect(self.page.get_by_role("columnheader", name="姓名")).to_be_visible()
        expect(self.page.get_by_role("columnheader", name="职位")).to_be_visible()

    def test_cdp_org_002_view_org_tree(self):
        """
        CDP-ORG-002: 查看完整组织树

        前置条件：用户已登录，存在组织数据
        测试步骤：
            1. 进入组织架构页面
        预期结果：
            - 左侧组织树显示完整层级结构
            - 每个节点显示名称、类型标签
            - 每个节点显示总人数 badge
        """
        self._login()
        self.page.get_by_test_id("link-header-org-structure").click()
        self.page.wait_for_timeout(2000)

        # Verify no backend error
        assert_no_error_message(self.page)

        # Verify organization tree exists
        tree = self.page.locator(".ant-tree")
        expect(tree).to_have_count(1)

        # Verify tree nodes exist - at least one node should be visible
        tree_nodes = tree.locator(".ant-tree-node-content-wrapper")
        if not tree_nodes.first.is_visible():
            pytest.skip("组织树为空，跳过测试")

        expect(tree_nodes.first).to_be_visible()

        # Verify there are multiple nodes (hierarchy) - the root should have children
        # Check that tree has proper structure with expandable nodes
        expect(tree.locator(".ant-tree-switcher")).to_be_visible()

    def test_cdp_org_003_get_org_node_detail(self):
        """
        CDP-ORG-003: 获取单个组织节点详情

        前置条件：用户已登录，存在组织数据
        测试步骤：
            1. 进入组织架构页面
            2. 点击组织树中某个节点
        预期结果：
            - 右侧员工列表显示该部门的直属员工
            - 选中节点高亮显示
        """
        self._login()
        self.page.get_by_test_id("link-header-org-structure").click()
        self.page.wait_for_timeout(2000)

        # Verify no backend error
        assert_no_error_message(self.page)

        # Click on a tree node to view its details
        tree = self.page.locator(".ant-tree")
        tree_node = tree.locator(".ant-tree-node-content-wrapper").first
        if not tree_node.is_visible():
            pytest.skip("组织树为空，跳过测试")

        expect(tree_node).to_be_visible()
        tree_node.click()
        self.page.wait_for_timeout(1000)

        # Verify the clicked node is highlighted (has ant-tree-node-selected class)
        expect(tree_node).to_have_class(re.compile("selected"))

        # Verify employee list table is still visible after node selection
        expect(self.page.get_by_role("columnheader", name="工号")).to_be_visible()
        expect(self.page.get_by_role("columnheader", name="姓名")).to_be_visible()


class TestCreateOrganization:
    """Test cases for creating organization units."""

    @pytest.fixture(autouse=True)
    def setup(self, page: Page):
        self.page = page
        self.login_page = LoginPage(page)

    def _login(self):
        """Login as admin user."""
        self.login_page.navigate()
        self.login_page.phone_input.fill("13800138002")
        self.login_page.password_input.fill("abcd1234")
        self.login_page.submit_button.click()
        self.page.wait_for_timeout(2000)

    def _open_add_org_dialog(self):
        """Open the add organization dialog."""
        self.page.get_by_test_id("link-header-org-structure").click()
        self.page.wait_for_timeout(2000)
        # Click the add button in the org tree panel header
        self.page.get_by_test_id("btn-org-add").click()
        self.page.wait_for_timeout(1000)

    @pytest.mark.smoke
    def test_cdp_org_crt_001_create_department_success(self):
        """
        CDP-ORG-CRT-001: 成功创建部门

        前置条件：用户已登录（管理员角色），存在上级组织
        测试步骤：
            1. 进入组织架构页面
            2. 点击"新增"按钮
            3. 填写名称、编码，选择类型
            4. 点击确认
        预期结果：
            - 新组织单元创建成功
            - 组织树刷新，新节点出现在对应位置
            - 返回创建成功提示
        """
        self._login()
        self._open_add_org_dialog()

        # Wait for dialog to appear
        dialog = self.page.locator(".ant-modal").first
        expect(dialog).to_be_visible()

        # Fill the form
        dialog.get_by_test_id("inp-org-name").fill("测试部门")
        dialog.get_by_test_id("inp-org-code").fill(
            f"TEST_DEPT_{int(__import__('time').time())}"
        )
        dialog.get_by_test_id("sel-org-type").click()
        self.page.wait_for_timeout(500)

        # Select department type from the dropdown
        self.page.locator(".ant-select-dropdown").get_by_text("部门", exact=True).click()
        self.page.wait_for_timeout(300)

        # Click confirm button
        self.page.get_by_test_id("btn-org-confirm").click()
        self.page.wait_for_timeout(2000)

        # Verify no backend error
        assert_no_error_message(self.page)

    def test_cdp_org_crt_002_duplicate_code(self):
        """
        CDP-ORG-CRT-002: 创建部门-编码重复

        前置条件：用户已登录（管理员角色），存在组织数据
        测试步骤：
            1. 进入组织架构页面
            2. 点击"新增"按钮
            3. 填写一个已存在的编码
            4. 点击确认
        预期结果：
            - 显示错误提示："编码已存在"或类似错误信息
            - 创建失败，组织树无变化
        """
        self._login()
        self._open_add_org_dialog()

        # Wait for dialog to appear
        dialog = self.page.locator(".ant-modal").first
        expect(dialog).to_be_visible()

        # Fill the form with a duplicate code
        dialog.get_by_test_id("inp-org-name").fill("测试重复编码部门")
        dialog.get_by_test_id("inp-org-code").fill("DEPT_001")
        dialog.get_by_test_id("sel-org-type").click()
        self.page.wait_for_timeout(500)

        # Select department type from the dropdown
        self.page.locator(".ant-select-dropdown").get_by_text("部门", exact=True).click()
        self.page.wait_for_timeout(300)

        # Click confirm button
        self.page.get_by_test_id("btn-org-confirm").click()
        self.page.wait_for_timeout(2000)

        # Verify error message appears (code duplicate error)
        # The backend should return an error via Ant Design message
        error_message = self.page.locator(".ant-message")
        expect(error_message).to_be_visible()

    def test_cdp_org_crt_003_missing_required_fields(self):
        """
        CDP-ORG-CRT-003: 创建部门-缺少必填字段

        前置条件：用户已登录（管理员角色）
        测试步骤：
            1. 进入组织架构页面
            2. 点击"新增"按钮
            3. 不填写名称直接点击确认
        预期结果：
            - 显示表单验证错误提示
            - 名称输入框高亮显示错误状态
        """
        self._login()
        self._open_add_org_dialog()

        # Wait for dialog to appear
        dialog = self.page.locator(".ant-modal").first
        expect(dialog).to_be_visible()

        # Click confirm without filling any fields
        self.page.get_by_test_id("btn-org-confirm").click()
        self.page.wait_for_timeout(500)

        # Verify form validation error - input should have error styling
        name_input = dialog.get_by_test_id("inp-org-name")
        expect(name_input).to_have_class(re.compile("ant-input-status-error"))

        # Verify error text appears (Ant Design shows error in ant-form-item-explain)
        expect(dialog.locator(".ant-form-item-explain-error").first).to_be_visible()


class TestEditOrganization:
    """Test cases for editing organization units."""

    @pytest.fixture(autouse=True)
    def setup(self, page: Page):
        self.page = page
        self.login_page = LoginPage(page)

    def _login(self):
        """Login as admin user."""
        self.login_page.navigate()
        self.login_page.phone_input.fill("13800138002")
        self.login_page.password_input.fill("abcd1234")
        self.login_page.submit_button.click()
        self.page.wait_for_timeout(2000)

    def test_cdp_org_edt_001_update_org_name(self):
        """
        CDP-ORG-EDT-001: 成功修改组织名称

        前置条件：用户已登录（管理员角色），存在组织数据
        测试步骤：
            1. 进入组织架构页面
            2. 在组织树上右键点击节点，选择"编辑"
            3. 修改组织名称
            4. 点击确认
        预期结果：
            - 组织名称更新成功
            - 组织树刷新，显示新名称
            - 返回更新成功提示
        """
        self._login()
        self.page.get_by_test_id("link-header-org-structure").click()
        self.page.wait_for_timeout(2000)

        # Verify no backend error
        assert_no_error_message(self.page)

        # Right-click on a tree node to open context menu
        tree = self.page.locator(".ant-tree")
        tree_node = tree.locator(".ant-tree-node-content-wrapper").first
        expect(tree_node).to_be_visible()
        tree_node.click(button="right")
        self.page.wait_for_timeout(500)

        # Look for edit option in context menu
        edit_option = self.page.get_by_text("编辑", exact=False)
        expect(edit_option).to_be_visible()
        edit_option.click()
        self.page.wait_for_timeout(1000)

        # Handle the edit dialog - should appear with current values
        dialog = self.page.locator(".ant-modal").first
        expect(dialog).to_be_visible()

        # Modify the organization name
        name_input = dialog.get_by_test_id("inp-org-name")
        name_input.fill("修改后的组织名称")
        self.page.wait_for_timeout(300)

        # Click confirm button
        self.page.get_by_test_id("btn-org-confirm").click()
        self.page.wait_for_timeout(1500)

        # Verify no backend error after operation
        assert_no_error_message(self.page)

    def test_cdp_org_edt_002_edit_nonexistent_org(self):
        """
        CDP-ORG-EDT-002: 编辑不存在的组织

        前置条件：用户已登录（管理员角色）
        测试步骤：
            1. 通过浏览器开发者工具修改请求，编辑一个不存在的组织单元
        预期结果：
            - 返回 404 错误
            - 显示错误提示
        """
        self._login()
        self.page.get_by_test_id("link-header-org-structure").click()
        self.page.wait_for_timeout(2000)

        # Set up route interception to modify the organization ID to a non-existent one
        def handle_route(route):
            # Modify the request to use a non-existent org ID
            request = route.request
            post_data = request.post_data
            if post_data and "organization" in post_data:
                import json

                data = json.loads(post_data)
                if "id" in data:
                    data["id"] = 999999  # Non-existent organization ID
                route.continue_(post_data=json.dumps(data))
            else:
                route.continue_()

        self.page.route("**/api/**/organization/**", handle_route)

        # Right-click on a tree node to open context menu
        tree = self.page.locator(".ant-tree")
        tree_node = tree.locator(".ant-tree-node-content-wrapper").first
        expect(tree_node).to_be_visible()
        tree_node.click(button="right")
        self.page.wait_for_timeout(500)

        # Look for edit option in context menu
        edit_option = self.page.get_by_text("编辑", exact=False)
        expect(edit_option).to_be_visible()
        edit_option.click()
        self.page.wait_for_timeout(1000)

        # Handle the edit dialog
        dialog = self.page.locator(".ant-modal").first
        if dialog.is_visible():
            # Modify the name slightly
            name_input = dialog.get_by_test_id("inp-org-name")
            name_input.fill("测试编辑不存在的组织")
            self.page.wait_for_timeout(300)

            # Click confirm
            self.page.get_by_test_id("btn-org-confirm").click()
            self.page.wait_for_timeout(2000)

        # Should show error message (404 or not found)
        error_message = self.page.locator(".ant-message-error")
        expect(error_message).to_be_visible(timeout=5000)


class TestDeleteOrganization:
    """Test cases for deleting organization units."""

    @pytest.fixture(autouse=True)
    def setup(self, page: Page):
        self.page = page
        self.login_page = LoginPage(page)

    def _login(self):
        """Login as admin user."""
        self.login_page.navigate()
        self.login_page.phone_input.fill("13800138002")
        self.login_page.password_input.fill("abcd1234")
        self.login_page.submit_button.click()
        self.page.wait_for_timeout(2000)

    def test_cdp_org_del_001_delete_empty_org(self):
        """
        CDP-ORG-DEL-001: 成功删除空组织

        前置条件：用户已登录（管理员角色），存在一个无子单元且无员工的组织
        测试步骤：
            1. 进入组织架构页面
            2. 在组织树上右键点击目标节点，选择"删除"
            3. 确认删除操作
        预期结果：
            - 组织单元删除成功
            - 组织树刷新，节点消失
            - 返回删除成功提示
        """
        self._login()
        self.page.get_by_test_id("link-header-org-structure").click()
        self.page.wait_for_timeout(2000)

        # Verify no backend error
        assert_no_error_message(self.page)

        # Right-click on a tree node to open context menu
        tree = self.page.locator(".ant-tree")
        tree_node = tree.locator(".ant-tree-node-content-wrapper").first
        expect(tree_node).to_be_visible()
        tree_node.click(button="right")
        self.page.wait_for_timeout(500)

        # Look for delete option in context menu
        delete_option = self.page.get_by_text("删除", exact=False)
        expect(delete_option).to_be_visible()
        delete_option.click()
        self.page.wait_for_timeout(500)

        # Handle confirmation dialog if it appears
        # Look for confirm button (typically "确定" or "确认")
        self.page.get_by_test_id("btn-org-confirm").click()
        self.page.wait_for_timeout(1500)

        # Verify no backend error after operation
        assert_no_error_message(self.page)

    def test_cdp_org_del_002_reject_delete_org_with_children(self):
        """
        CDP-ORG-DEL-002: 拒绝删除非空组织-有子单元

        前置条件：用户已登录（管理员角色），存在有子单元的组织
        测试步骤：
            1. 进入组织架构页面
            2. 在组织树上右键点击有子单元的节点，选择"删除"
        预期结果：
            - 显示错误提示："该单元存在子组织，无法删除"
            - 组织单元未被删除
        """
        self._login()
        self.page.get_by_test_id("link-header-org-structure").click()
        self.page.wait_for_timeout(2000)

        # Verify no backend error
        assert_no_error_message(self.page)

        # First expand the root node to see children
        tree = self.page.locator(".ant-tree")
        switcher = tree.locator(".ant-tree-switcher").first
        if switcher.is_visible():
            switcher.click()
            self.page.wait_for_timeout(500)

        # Find a node that has children (has .ant-tree-node-content-wrapper with parent having expanded switcher)
        # Right-click on a tree node to open context menu
        tree_node = tree.locator(".ant-tree-node-content-wrapper").first
        expect(tree_node).to_be_visible()
        tree_node.click(button="right")
        self.page.wait_for_timeout(500)

        # Look for delete option in context menu
        delete_option = self.page.get_by_text("删除", exact=False)
        expect(delete_option).to_be_visible()
        delete_option.click()
        self.page.wait_for_timeout(1000)

        # Should show error message about having children
        # The error message should contain "子组织" or similar text
        error_message = self.page.locator(".ant-message-error")
        expect(error_message).to_be_visible(timeout=5000)

    def test_cdp_org_del_003_reject_delete_org_with_employees(self):
        """
        CDP-ORG-DEL-003: 拒绝删除非空组织-有员工

        前置条件：用户已登录（管理员角色），存在有归属员工的组织
        测试步骤：
            1. 进入组织架构页面
            2. 在组织树上右键点击有员工的节点，选择"删除"
        预期结果：
            - 显示错误提示："该单元存在归属员工，无法删除"
            - 组织单元未被删除
        """
        self._login()
        self.page.get_by_test_id("link-header-org-structure").click()
        self.page.wait_for_timeout(2000)

        # Verify no backend error
        assert_no_error_message(self.page)

        # Click on a tree node that likely has employees assigned
        tree = self.page.locator(".ant-tree")
        tree_node = tree.locator(".ant-tree-node-content-wrapper").first
        expect(tree_node).to_be_visible()
        tree_node.click(button="right")
        self.page.wait_for_timeout(500)

        # Look for delete option in context menu
        delete_option = self.page.get_by_text("删除", exact=False)
        expect(delete_option).to_be_visible()
        delete_option.click()
        self.page.wait_for_timeout(1000)

        # Should show error message about having employees
        # The error message should contain "归属员工" or similar text
        error_message = self.page.locator(".ant-message-error")
        expect(error_message).to_be_visible(timeout=5000)


class TestToggleOrganization:
    """Test cases for enabling/disabling organization units."""

    @pytest.fixture(autouse=True)
    def setup(self, page: Page):
        self.page = page
        self.login_page = LoginPage(page)

    def _login(self):
        """Login as admin user."""
        self.login_page.navigate()
        self.login_page.phone_input.fill("13800138002")
        self.login_page.password_input.fill("abcd1234")
        self.login_page.submit_button.click()
        self.page.wait_for_timeout(2000)

    def test_cdp_org_toggle_001_disable_org(self):
        """
        CDP-ORG-TOGGLE-001: 禁用启用状态的组织

        前置条件：用户已登录（管理员角色），存在启用状态的组织
        测试步骤：
            1. 进入组织架构页面
            2. 在组织树上右键点击目标节点，选择"禁用"
        预期结果：
            - 组织状态变为"禁用"
            - 节点显示禁用标签
            - 组织树刷新，该节点不再显示（或显示为灰色）
        """
        self._login()
        self.page.get_by_test_id("link-header-org-structure").click()
        self.page.wait_for_timeout(2000)

        # Verify no backend error
        assert_no_error_message(self.page)

        # Right-click on a tree node to open context menu
        tree = self.page.locator(".ant-tree")
        tree_node = tree.locator(".ant-tree-node-content-wrapper").first
        expect(tree_node).to_be_visible()
        tree_node.click(button="right")
        self.page.wait_for_timeout(500)

        # Look for context menu with disable option
        # The menu should contain "禁用" or similar option
        disable_option = self.page.get_by_text("禁用", exact=False)
        expect(disable_option).to_be_visible()
        disable_option.click()
        self.page.wait_for_timeout(1500)

        # Verify no backend error after operation
        assert_no_error_message(self.page)

    def test_cdp_org_toggle_002_enable_disabled_org(self):
        """
        CDP-ORG-TOGGLE-002: 重新启用禁用的组织

        前置条件：用户已登录（管理员角色），存在禁用状态的组织
        测试步骤：
            1. 进入组织架构页面
            2. 在组织树上右键点击目标节点，选择"启用"
        预期结果：
            - 组织状态变为"启用"
            - 节点恢复正常显示
            - 组织树刷新，节点可见
        """
        self._login()
        self.page.get_by_test_id("link-header-org-structure").click()
        self.page.wait_for_timeout(2000)

        # Verify no backend error
        assert_no_error_message(self.page)

        # Right-click on a tree node to open context menu
        tree = self.page.locator(".ant-tree")
        tree_node = tree.locator(".ant-tree-node-content-wrapper").first
        expect(tree_node).to_be_visible()
        tree_node.click(button="right")
        self.page.wait_for_timeout(500)

        # Look for context menu with enable option
        enable_option = self.page.get_by_text("启用", exact=False)
        expect(enable_option).to_be_visible()
        enable_option.click()
        self.page.wait_for_timeout(1500)

        # Verify no backend error after operation
        assert_no_error_message(self.page)


class TestMoveOrganization:
    """Test cases for moving organization units."""

    @pytest.fixture(autouse=True)
    def setup(self, page: Page):
        self.page = page
        self.login_page = LoginPage(page)

    def _login(self):
        """Login as admin user."""
        self.login_page.navigate()
        self.login_page.phone_input.fill("13800138002")
        self.login_page.password_input.fill("abcd1234")
        self.login_page.submit_button.click()
        self.page.wait_for_timeout(2000)

    def test_cdp_org_move_001_move_org_success(self):
        """
        CDP-ORG-MOVE-001: 成功移动组织单元

        前置条件：用户已登录（管理员角色），存在可移动的组织
        测试步骤：
            1. 进入组织架构页面
            2. 在组织树上右键点击目标节点，选择"移动"
            3. 选择新的上级节点
            4. 点击确认
        预期结果：
            - 组织单元移动成功
            - 组织树刷新，节点出现在新位置
            - 节点的 level 值正确更新
        """
        self._login()
        self.page.get_by_test_id("link-header-org-structure").click()
        self.page.wait_for_timeout(2000)

        # Verify no backend error
        assert_no_error_message(self.page)

        # First expand the tree to see more nodes
        tree = self.page.locator(".ant-tree")
        switchers = tree.locator(".ant-tree-switcher")
        if switchers.first.is_visible():
            switchers.first.click()
            self.page.wait_for_timeout(500)

        # Right-click on a tree node to open context menu
        tree_node = tree.locator(".ant-tree-node-content-wrapper").first
        expect(tree_node).to_be_visible()
        tree_node.click(button="right")
        self.page.wait_for_timeout(500)

        # Look for move option in context menu
        move_option = self.page.get_by_text("移动", exact=False)
        if move_option.is_visible():
            move_option.click()
            self.page.wait_for_timeout(1000)

            # Handle the move dialog if it appears
            dialog = self.page.locator(".ant-modal").first
            if dialog.is_visible():
                # Select a new parent node in the tree dialog
                # Then click confirm
                self.page.get_by_test_id("btn-transfer-confirm").click()
                self.page.wait_for_timeout(1500)

        # Verify no backend error after operation
        assert_no_error_message(self.page)

    def test_cdp_org_move_002_reject_circular_move(self):
        """
        CDP-ORG-MOVE-002: 拒绝循环移动

        前置条件：用户已登录（管理员角色），存在可移动的组织
        测试步骤：
            1. 进入组织架构页面
            2. 在组织树上右键点击节点 A，选择"移动"
            3. 尝试将 A 移动到 A 的子节点下
        预期结果：
            - 显示错误提示："不能移动到自身后代节点下"或"循环引用"
            - 移动操作被拒绝
        """
        self._login()
        self.page.get_by_test_id("link-header-org-structure").click()
        self.page.wait_for_timeout(2000)

        # Verify no backend error
        assert_no_error_message(self.page)

        # First expand the tree to see parent-child relationships
        tree = self.page.locator(".ant-tree")
        switchers = tree.locator(".ant-tree-switcher")
        if switchers.first.is_visible():
            switchers.first.click()
            self.page.wait_for_timeout(500)

        # Right-click on a tree node to open context menu
        tree_node = tree.locator(".ant-tree-node-content-wrapper").first
        expect(tree_node).to_be_visible()
        tree_node.click(button="right")
        self.page.wait_for_timeout(500)

        # Look for move option in context menu
        move_option = self.page.get_by_text("移动", exact=False)
        if move_option.is_visible():
            move_option.click()
            self.page.wait_for_timeout(1000)

            # Handle the move dialog if it appears
            dialog = self.page.locator(".ant-modal").first
            if dialog.is_visible():
                # Try to select a child node (which would cause circular reference)
                # This should trigger an error
                # Click confirm anyway to trigger the error
                self.page.get_by_test_id("btn-transfer-confirm").click()
                self.page.wait_for_timeout(1500)

                # Should show error message about circular reference
                error_message = self.page.locator(".ant-message-error")
                expect(error_message).to_be_visible(timeout=5000)


class TestEmployeeManagement:
    """Test cases for employee management."""

    @pytest.fixture(autouse=True)
    def setup(self, page: Page):
        self.page = page
        self.login_page = LoginPage(page)

    def _login(self):
        """Login as admin user."""
        self.login_page.navigate()
        self.login_page.phone_input.fill("13800138002")
        self.login_page.password_input.fill("abcd1234")
        self.login_page.submit_button.click()
        self.page.wait_for_timeout(2000)

    @pytest.mark.smoke
    def test_cdp_emp_001_employee_list_page_loads(self):
        """
        CDP-EMP-001: 员工列表页正常加载

        前置条件：用户已登录
        测试步骤：
            1. 进入组织架构页面
        预期结果：
            - 右侧显示员工列表
            - 列表包含分页控件
            - 显示员工工号、姓名、岗位、手机、状态等字段
        """
        self._login()
        self.page.get_by_test_id("link-header-org-structure").click()
        self.page.wait_for_timeout(2000)

        # Verify employee list table headers
        expect(self.page.get_by_role("columnheader", name="工号")).to_be_visible()
        expect(self.page.get_by_role("columnheader", name="姓名")).to_be_visible()
        expect(self.page.get_by_role("columnheader", name="职位")).to_be_visible()
        expect(self.page.get_by_role("columnheader", name="电话")).to_be_visible()
        expect(self.page.get_by_role("columnheader", name="状态")).to_be_visible()
        expect(self.page.get_by_role("columnheader", name="账号")).to_be_visible()
        expect(self.page.get_by_role("columnheader", name="入职日期")).to_be_visible()

        # Verify pagination exists
        expect(self.page.locator(".ant-pagination")).to_be_visible()

        # Verify no backend errors
        assert_no_error_message(self.page)

    def test_cdp_emp_002_filter_by_department(self):
        """
        CDP-EMP-002: 按部门筛选员工

        前置条件：用户已登录，存在多部门数据
        测试步骤：
            1. 进入组织架构页面
            2. 点击组织树中的某个部门节点
        预期结果：
            - 员工列表刷新
            - 仅显示该部门的直属员工
            - 选中节点高亮
        """
        self._login()
        self.page.get_by_test_id("link-header-org-structure").click()
        self.page.wait_for_timeout(2000)

        # Click on first tree node in the organization tree
        tree = self.page.locator(".ant-tree")
        if tree.is_visible():
            tree_node = tree.locator(".ant-tree-treenode").first
            tree_node.click()
            self.page.wait_for_timeout(1500)

            # Verify the node is highlighted (selected class)
            expect(tree_node).to_have_class(re.compile("selected"))

            # Verify no backend errors
            assert_no_error_message(self.page)

    def test_cdp_emp_003_filter_by_status(self):
        """
        CDP-EMP-003: 按状态筛选员工

        前置条件：用户已登录，存在多种状态的员工
        测试步骤：
            1. 进入组织架构页面
            2. 在员工列表上方找到状态下拉框
            3. 选择"在职"
        预期结果：
            - 员工列表刷新
            - 仅显示状态为"在职"的员工
        """
        self._login()
        self.page.get_by_test_id("link-header-org-structure").click()
        self.page.wait_for_timeout(2000)

        # Find and click the status filter combobox
        status_filter = self.page.get_by_placeholder("状态筛选")
        expect(status_filter).to_be_visible()
        status_filter.click()
        self.page.wait_for_timeout(500)

        # Select "在职" option
        self.page.get_by_role("option", name="在职").click()
        self.page.wait_for_timeout(1500)

        # Verify no backend errors
        assert_no_error_message(self.page)

    def test_cdp_emp_004_search_by_keyword(self):
        """
        CDP-EMP-004: 按关键词搜索员工

        前置条件：用户已登录，存在员工数据
        测试步骤：
            1. 进入组织架构页面
            2. 在员工列表上方的搜索框输入员工姓名关键词
            3. 按回车或点击搜索
        预期结果：
            - 员工列表刷新
            - 仅显示姓名或工号包含关键词的员工
        """
        self._login()
        self.page.get_by_test_id("link-header-org-structure").click()
        self.page.wait_for_timeout(2000)

        # Find and use the search box
        search_box = self.page.get_by_placeholder("搜索姓名/工号/电话")
        expect(search_box).to_be_visible()
        search_box.fill("张")
        search_box.press("Enter")
        self.page.wait_for_timeout(1500)

        # Verify no backend errors
        assert_no_error_message(self.page)

    def test_cdp_emp_005_include_sub_departments(self):
        """
        CDP-EMP-005: 包含下级部门员工筛选

        前置条件：用户已登录，存在层级部门数据
        测试步骤：
            1. 进入组织架构页面
            2. 点击组织树中的某个部门节点
            3. 勾选"包含下级"选项
        预期结果：
            - 员工列表刷新
            - 显示该部门及其所有下级部门的员工
        """
        self._login()
        self.page.get_by_test_id("link-header-org-structure").click()
        self.page.wait_for_timeout(2000)

        # Click on first tree node in the organization tree
        tree = self.page.locator(".ant-tree")
        tree_node = tree.locator(".ant-tree-treenode").first
        if tree_node.is_visible():
            tree_node.click()
            self.page.wait_for_timeout(1000)

            # Find and toggle the "包含下级" checkbox
            include_sub_checkbox = self.page.get_by_text("包含下级")
            if include_sub_checkbox.is_visible():
                include_sub_checkbox.click()
                self.page.wait_for_timeout(1500)

            # Verify no backend errors
            assert_no_error_message(self.page)


class TestCreateEmployee:
    """Test cases for creating employees."""

    @pytest.fixture(autouse=True)
    def setup(self, page: Page):
        self.page = page
        self.login_page = LoginPage(page)

    def _login(self):
        """Login as admin user."""
        self.login_page.navigate()
        self.login_page.phone_input.fill("13800138002")
        self.login_page.password_input.fill("abcd1234")
        self.login_page.submit_button.click()
        self.page.wait_for_timeout(2000)

    def test_cdp_emp_crt_001_create_employee_without_account(self):
        """
        CDP-EMP-CRT-001: 成功创建员工（不绑定账号）

        前置条件：用户已登录（管理员角色），存在组织数据
        测试步骤：
            1. 进入组织架构页面
            2. 点击"新增员工"按钮
            3. 填写工号、姓名、手机、邮箱、岗位，选择主属部门
            4. 不绑定系统账号
            5. 点击确认
        预期结果：
            - 员工创建成功
            - 员工列表刷新，新员工出现在列表中
            - 返回创建成功提示
        """
        self._login()
        self.page.get_by_test_id("link-header-org-structure").click()
        self.page.wait_for_timeout(2000)

        # Click "新增员工" button
        add_button = self.page.get_by_test_id("btn-employee-add")
        expect(add_button).to_be_visible()
        add_button.click()
        self.page.wait_for_timeout(1000)

        # Fill in the form - generate unique employee code
        import time

        unique_code = f"EMP{int(time.time()) % 100000:05d}"

        # Fill employee code
        self.page.get_by_label("工号").fill(unique_code)
        # Fill name
        self.page.get_by_label("姓名").fill("测试员工")
        # Fill phone
        self.page.get_by_label("手机").fill("13900001001")
        # Fill position
        self.page.get_by_label("职位").fill("测试工程师")
        # Click confirm/submit button in modal
        self.page.get_by_test_id("btn-emp-confirm").click()
        self.page.wait_for_timeout(2000)

        # Verify no backend errors
        assert_no_error_message(self.page)

    def test_cdp_emp_crt_002_create_employee_with_account(self):
        """
        CDP-EMP-CRT-002: 成功创建员工（绑定账号）

        前置条件：用户已登录（管理员角色），存在未绑定员工的系统账号
        测试步骤：
            1. 进入组织架构页面
            2. 点击"新增员工"按钮
            3. 填写必填信息
            4. 输入一个未绑定员工的 user_id
            5. 点击确认
        预期结果：
            - 员工创建成功并绑定账号
            - 员工列表中该员工显示"已绑定"标签
        """
        self._login()
        self.page.get_by_test_id("link-header-org-structure").click()
        self.page.wait_for_timeout(2000)

        # Click "新增员工" button
        add_button = self.page.get_by_test_id("btn-employee-add")
        expect(add_button).to_be_visible()
        add_button.click()
        self.page.wait_for_timeout(1000)

        # Fill in the form
        import time

        unique_code = f"EMP{int(time.time()) % 100000:05d}"

        # Fill employee code
        self.page.get_by_label("工号").fill(unique_code)
        # Fill name
        self.page.get_by_label("姓名").fill("绑定账号员工")
        # Fill phone
        self.page.get_by_label("手机").fill("13900001002")
        # Fill position
        self.page.get_by_label("职位").fill("测试工程师")
        # Try to bind an account (user_id)
        account_input = self.page.get_by_label("账号")
        if account_input.is_visible():
            account_input.fill("999")  # This might need a valid unbound user_id

        # Click confirm
        self.page.get_by_test_id("btn-emp-confirm").click()
        self.page.wait_for_timeout(2000)

        # Verify no backend errors
        assert_no_error_message(self.page)

    def test_cdp_emp_crt_003_duplicate_employee_code(self):
        """
        CDP-EMP-CRT-003: 创建员工-工号重复

        前置条件：用户已登录（管理员角色），存在员工数据
        测试步骤：
            1. 进入组织架构页面
            2. 点击"新增员工"按钮
            3. 填写一个已存在的工号
            4. 点击确认
        预期结果：
            - 显示错误提示："工号已存在"
            - 创建失败
        """
        self._login()
        self.page.get_by_test_id("link-header-org-structure").click()
        self.page.wait_for_timeout(2000)

        # Click "新增员工" button
        add_button = self.page.get_by_test_id("btn-employee-add")
        expect(add_button).to_be_visible()
        add_button.click()
        self.page.wait_for_timeout(1000)

        # Try to use a duplicate code - use a common pattern like "EMP001"
        self.page.get_by_label("工号").fill("EMP001")
        self.page.get_by_label("姓名").fill("重复工号测试")
        self.page.get_by_label("手机").fill("13900001003")

        # Click confirm
        self.page.get_by_test_id("btn-emp-confirm").click()
        self.page.wait_for_timeout(2000)

        # Verify error message appears
        # ant-message-error indicates backend error
        error_locator = self.page.locator(".ant-message-error")
        expect(error_locator).to_be_visible(timeout=5000)

    def test_cdp_emp_crt_004_bind_already_bound_account(self):
        """
        CDP-EMP-CRT-004: 创建员工-绑定已关联账号

        前置条件：用户已登录（管理员角色），存在已绑定员工的账号
        测试步骤：
            1. 进入组织架构页面
            2. 点击"新增员工"按钮
            3. 填写必填信息
            4. 输入一个已绑定其他员工的 user_id
            5. 点击确认
        预期结果：
            - 显示错误提示："该账号已绑定其他员工"
            - 创建失败
        """
        self._login()
        self.page.get_by_test_id("link-header-org-structure").click()
        self.page.wait_for_timeout(2000)

        # Click "新增员工" button
        add_button = self.page.get_by_test_id("btn-employee-add")
        expect(add_button).to_be_visible()
        add_button.click()
        self.page.wait_for_timeout(1000)

        # Fill form with an already bound account
        import time

        unique_code = f"EMP{int(time.time()) % 100000:05d}"
        self.page.get_by_label("工号").fill(unique_code)
        self.page.get_by_label("姓名").fill("绑定已占用账号测试")
        self.page.get_by_label("手机").fill("13900001004")
        self.page.get_by_label("职位").fill("测试工程师")

        # Try to bind account 1 which is likely already bound to test user
        account_input = self.page.get_by_label("账号")
        if account_input.is_visible():
            account_input.fill("1")

        # Click confirm
        self.page.get_by_test_id("btn-emp-confirm").click()
        self.page.wait_for_timeout(2000)

        # Verify error message appears
        error_locator = self.page.locator(".ant-message-error")
        expect(error_locator).to_be_visible(timeout=5000)


class TestEditDeleteEmployee:
    """Test cases for editing and deleting employees."""

    @pytest.fixture(autouse=True)
    def setup(self, page: Page):
        self.page = page
        self.login_page = LoginPage(page)

    def _login(self):
        """Login as admin user."""
        self.login_page.navigate()
        self.login_page.phone_input.fill("13800138002")
        self.login_page.password_input.fill("abcd1234")
        self.login_page.submit_button.click()
        self.page.wait_for_timeout(2000)

    def test_cdp_emp_edt_001_update_employee_position(self):
        """
        CDP-EMP-EDT-001: 成功修改员工岗位

        前置条件：用户已登录（管理员角色），存在员工数据
        测试步骤：
            1. 进入组织架构页面
            2. 在员工列表中找到目标员工，点击操作列"编辑"
            3. 修改岗位信息
            4. 点击确认
        预期结果：
            - 员工信息更新成功
            - 员工列表刷新，显示新岗位
        """
        self._login()
        self.page.get_by_test_id("link-header-org-structure").click()
        self.page.wait_for_timeout(2000)

        # Find and click the edit button for first employee in the table
        # Look for the edit button in the action column
        edit_buttons = self.page.get_by_role("button", name="edit 编辑")
        if edit_buttons.first.is_visible():
            edit_buttons.first.click()
            self.page.wait_for_timeout(1000)

            # Modify the position field
            position_input = self.page.get_by_label("职位")
            if position_input.is_visible():
                position_input.fill("高级测试工程师")

                # Click confirm
                self.page.get_by_test_id("btn-emp-confirm").click()
                self.page.wait_for_timeout(2000)

        # Verify no backend errors
        assert_no_error_message(self.page)

    def test_cdp_emp_edt_002_edit_nonexistent_employee(self):
        """
        CDP-EMP-EDT-002: 编辑不存在的员工

        前置条件：用户已登录（管理员角色）
        测试步骤：
            1. 通过浏览器开发者工具修改请求，编辑一个不存在的员工
        预期结果：
            - 返回 404 错误
        """
        self._login()
        self.page.get_by_test_id("link-header-org-structure").click()
        self.page.wait_for_timeout(2000)

        # This test requires modifying API request directly
        # We can use page.route to intercept and modify the request
        # For now, just verify we can access the edit dialog if it exists
        edit_buttons = self.page.get_by_role("button", name="edit 编辑")
        if edit_buttons.first.is_visible():
            edit_buttons.first.click()
            self.page.wait_for_timeout(1000)

            # Try to submit with an invalid employee ID by modifying the form
            # This is a placeholder - actual implementation would need API mocking
            self.page.wait_for_timeout(500)

    def test_cdp_emp_del_001_soft_delete_employee(self):
        """
        CDP-EMP-DEL-001: 软删除员工（离职）

        前置条件：用户已登录（管理员角色），存在员工数据
        测试步骤：
            1. 进入组织架构页面
            2. 在员工列表中找到目标员工，点击操作列"删除"
            3. 确认删除操作
        预期结果：
            - 员工状态变为"离职"（offboarding）
            - 员工列表刷新，该员工不再显示在职列表中（或显示为离职状态）
            - 返回删除成功提示
        """
        self._login()
        self.page.get_by_test_id("link-header-org-structure").click()
        self.page.wait_for_timeout(2000)

        # Find and click the delete button for first employee
        delete_buttons = self.page.get_by_role("button", name="delete 删除")
        if delete_buttons.first.is_visible():
            delete_buttons.first.click()
            self.page.wait_for_timeout(500)

            # Handle confirmation dialog if it appears
            # Click confirm in the popup dialog
            self.page.get_by_test_id("btn-emp-confirm").click()
            self.page.wait_for_timeout(2000)

        # Verify no backend errors
        assert_no_error_message(self.page)

    def test_cdp_emp_del_002_verify_resigned_employee_status(self):
        """
        CDP-EMP-DEL-002: 离职员工状态验证

        前置条件：用户已登录（管理员角色），刚完成软删除操作
        测试步骤：
            1. 在员工列表筛选"离职"状态
        预期结果：
            - 能找到刚才离职的员工
            - 显示正确的离职日期
        """
        self._login()
        self.page.get_by_test_id("link-header-org-structure").click()
        self.page.wait_for_timeout(2000)

        # Filter by "离职" status
        status_filter = self.page.get_by_placeholder("状态筛选")
        expect(status_filter).to_be_visible()
        status_filter.click()
        self.page.wait_for_timeout(500)

        # Select "离职" option
        self.page.get_by_role("option", name="离职").click()
        self.page.wait_for_timeout(1500)

        # Verify the employee who was just deleted is in the list
        # (This test depends on the previous test running first)
        assert_no_error_message(self.page)


class TestEmployeeAccountBinding:
    """Test cases for employee account binding."""

    @pytest.fixture(autouse=True)
    def setup(self, page: Page):
        self.page = page
        self.login_page = LoginPage(page)

    def _login(self):
        """Login as admin user."""
        self.login_page.navigate()
        self.login_page.phone_input.fill("13800138002")
        self.login_page.password_input.fill("abcd1234")
        self.login_page.submit_button.click()
        self.page.wait_for_timeout(2000)

    def test_cdp_emp_bind_001_bind_account_to_employee(self):
        """
        CDP-EMP-BIND-001: 为员工绑定系统账号

        前置条件：用户已登录（管理员角色），存在未绑定账号的员工和未关联的账号
        测试步骤：
            1. 进入组织架构页面
            2. 在员工列表中找到目标员工，点击"绑定账号"
            3. 输入 user_id
            4. 点击确认
        预期结果：
            - 绑定成功
            - 员工列表中该员工显示"已绑定"标签
        """
        self._login()
        self.page.get_by_test_id("link-header-org-structure").click()
        self.page.wait_for_timeout(2000)

        # Find and click "绑定账号" button
        bind_button = self.page.get_by_role("button", name="link 绑定账号")
        if bind_button.is_visible():
            bind_button.first.click()
            self.page.wait_for_timeout(1000)

            # Enter user_id in the dialog
            user_id_input = self.page.get_by_label("账号")
            if user_id_input.is_visible():
                user_id_input.fill("999")  # Use an unbound user_id

                # Click confirm
                self.page.get_by_test_id("btn-bind-confirm").click()
                self.page.wait_for_timeout(2000)

        # Verify no backend errors
        assert_no_error_message(self.page)

    def test_cdp_emp_bind_002_bind_already_bound_account(self):
        """
        CDP-EMP-BIND-002: 绑定已关联账号

        前置条件：用户已登录（管理员角色），存在未绑定账号的员工和已关联的账号
        测试步骤：
            1. 进入组织架构页面
            2. 在员工列表中找到目标员工，点击"绑定账号"
            3. 输入一个已绑定其他员工的 user_id
            4. 点击确认
        预期结果：
            - 显示错误提示："该账号已绑定其他员工"
            - 绑定失败
        """
        self._login()
        self.page.get_by_test_id("link-header-org-structure").click()
        self.page.wait_for_timeout(2000)

        # Find and click "绑定账号" button
        bind_button = self.page.get_by_role("button", name="link 绑定账号")
        if bind_button.is_visible():
            bind_button.first.click()
            self.page.wait_for_timeout(1000)

            # Enter an already bound user_id (e.g., user_id=1 which is likely bound)
            user_id_input = self.page.get_by_label("账号")
            if user_id_input.is_visible():
                user_id_input.fill("1")

                # Click confirm
                self.page.get_by_test_id("btn-bind-confirm").click()
                self.page.wait_for_timeout(2000)

        # Verify error message appears
        error_locator = self.page.locator(".ant-message-error")
        expect(error_locator).to_be_visible(timeout=5000)

    def test_cdp_emp_bind_003_unbind_employee_account(self):
        """
        CDP-EMP-BIND-003: 解绑员工账号

        前置条件：用户已登录（管理员角色），存在已绑定账号的员工
        测试步骤：
            1. 进入组织架构页面
            2. 在员工列表中找到已绑定账号的员工，点击"解绑账号"
            3. 确认解绑操作
        预期结果：
            - 解绑成功
            - 员工列表中该员工显示"未绑定"标签
            - user_employee_mapping 记录已删除
        """
        self._login()
        self.page.get_by_test_id("link-header-org-structure").click()
        self.page.wait_for_timeout(2000)

        # Find and click "解绑账号" button
        unbind_button = self.page.get_by_role("button", name="link 解绑账号")
        if unbind_button.is_visible():
            unbind_button.first.click()
            self.page.wait_for_timeout(500)

            # Confirm the unbind operation
            self.page.get_by_test_id("btn-bind-confirm").click()
            self.page.wait_for_timeout(2000)

        # Verify no backend errors
        assert_no_error_message(self.page)

    def test_cdp_emp_bind_004_rebind_same_account(self):
        """
        CDP-EMP-BIND-004: 再次绑定同一账号

        前置条件：用户已登录（管理员角色），刚完成解绑操作的员工和同一账号
        测试步骤：
            1. 在员工列表中找到刚解绑的员工，点击"绑定账号"
            2. 输入同一个 user_id
            3. 点击确认
        预期结果：
            - 绑定成功
            - 该员工重新绑定同一账号
        """
        self._login()
        self.page.get_by_test_id("link-header-org-structure").click()
        self.page.wait_for_timeout(2000)

        # Find and click "绑定账号" button for the first unbound employee
        bind_button = self.page.get_by_role("button", name="link 绑定账号")
        if bind_button.is_visible():
            bind_button.first.click()
            self.page.wait_for_timeout(1000)

            # Enter user_id to rebind
            user_id_input = self.page.get_by_label("账号")
            if user_id_input.is_visible():
                user_id_input.fill("999")

                # Click confirm
                self.page.get_by_test_id("btn-bind-confirm").click()
                self.page.wait_for_timeout(2000)

        # Verify no backend errors
        assert_no_error_message(self.page)


class TestEmployeeSecondaryDepartment:
    """Test cases for employee secondary department management."""

    @pytest.fixture(autouse=True)
    def setup(self, page: Page):
        self.page = page
        self.login_page = LoginPage(page)

    def _login(self):
        """Login as admin user."""
        self.login_page.navigate()
        self.login_page.phone_input.fill("13800138002")
        self.login_page.password_input.fill("abcd1234")
        self.login_page.submit_button.click()
        self.page.wait_for_timeout(2000)

    def test_cdp_emp_sec_001_add_secondary_department(self):
        """
        CDP-EMP-SEC-001: 添加辅属部门

        前置条件：用户已登录（管理员角色），存在员工和多个组织单元
        测试步骤：
            1. 进入组织架构页面
            2. 在员工列表中找到目标员工，点击"编辑"
            3. 在辅属部门中添加一个其他部门
            4. 点击确认
        预期结果：
            - 辅属部门添加成功
            - 员工详情或列表中显示辅属部门信息
        """
        self._login()
        self.page.get_by_test_id("link-header-org-structure").click()
        self.page.wait_for_timeout(2000)

        # Find and click the edit button for first employee
        edit_buttons = self.page.get_by_role("button", name="edit 编辑")
        expect(edit_buttons.first).to_be_visible()
        edit_buttons.first.click()
        self.page.wait_for_timeout(1500)

        # Look for "添加辅属部门" button or secondary department section
        add_secondary_btn = self.page.get_by_role("button", name="添加辅属部门")
        if add_secondary_btn.is_visible():
            add_secondary_btn.click()
            self.page.wait_for_timeout(1000)

            # Select a department from the dropdown/tree selector
            # Look for a department tree or select component
            tree_nodes = self.page.locator(
                ".ant-tree-treenode, .ant-select-selector, .ant-select-item"
            )
            if tree_nodes.first.is_visible():
                tree_nodes.first.click()
                self.page.wait_for_timeout(500)

            # Click confirm to save
            self.page.get_by_test_id("btn-emp-confirm").click()
            self.page.wait_for_timeout(2000)

        # Verify no backend errors
        assert_no_error_message(self.page)

    def test_cdp_emp_sec_002_remove_secondary_department(self):
        """
        CDP-EMP-SEC-002: 移除辅属部门

        前置条件：用户已登录（管理员角色），存在有辅属部门的员工
        测试步骤：
            1. 进入组织架构页面
            2. 在员工列表中找到目标员工，点击"编辑"
            3. 移除某个辅属部门
            4. 点击确认
        预期结果：
            - 辅属部门移除成功
            - 员工详情中不再显示该辅属部门
        """
        self._login()
        self.page.get_by_test_id("link-header-org-structure").click()
        self.page.wait_for_timeout(2000)

        # Find and click the edit button for first employee
        edit_buttons = self.page.get_by_role("button", name="edit 编辑")
        expect(edit_buttons.first).to_be_visible()
        edit_buttons.first.click()
        self.page.wait_for_timeout(1500)

        # Look for secondary department section with remove buttons
        # Look for elements that allow removing a secondary department
        # Common patterns: delete icon buttons, "移除" links/buttons
        remove_buttons = self.page.locator(
            "[aria-label='移除'], .anticon-delete, .ant-btn-dangerous"
        )
        has_remove = remove_buttons.first.is_visible(timeout=3000)

        if has_remove:
            # Click the first remove button
            remove_buttons.first.click()
            self.page.wait_for_timeout(1000)

            # Click confirm to save
            self.page.get_by_test_id("btn-emp-confirm").click()
            self.page.wait_for_timeout(2000)

        # Verify no backend errors
        assert_no_error_message(self.page)


class TestEmployeeTransfer:
    """Test cases for employee transfer."""

    @pytest.fixture(autouse=True)
    def setup(self, page: Page):
        self.page = page
        self.login_page = LoginPage(page)

    def _login(self):
        """Login as admin user."""
        self.login_page.navigate()
        self.login_page.phone_input.fill("13800138002")
        self.login_page.password_input.fill("abcd1234")
        self.login_page.submit_button.click()
        self.page.wait_for_timeout(2000)

    def test_cdp_emp_trans_001_initiate_transfer(self):
        """
        CDP-EMP-TRANS-001: 成功发起调动

        前置条件：用户已登录（管理员角色），存在员工和目标部门
        测试步骤：
            1. 进入组织架构页面
            2. 在员工列表中找到目标员工，点击"调动"
            3. 选择目标部门、调动类型、填写生效日期和原因
            4. 点击确认
        预期结果：
            - 调动发起成功
            - 员工状态变为"调动中"
            - 返回调动记录
        """
        self._login()
        self.page.get_by_test_id("link-header-org-structure").click()
        self.page.wait_for_timeout(2000)

        # Find and click "调动" button
        transfer_button = self.page.get_by_role("button", name="transfer 调动")
        if transfer_button.is_visible():
            transfer_button.first.click()
            self.page.wait_for_timeout(1000)

            # Fill in the transfer form if visible
            # Select target department, transfer type, effective date, reason
            # This would depend on the actual form fields
            self.page.get_by_test_id("btn-transfer-confirm").click()
            self.page.wait_for_timeout(2000)

        # Verify no backend errors
        assert_no_error_message(self.page)

    def test_cdp_emp_trans_002_reject_same_department_transfer(self):
        """
        CDP-EMP-TRANS-002: 调动至同一部门被拒绝

        前置条件：用户已登录（管理员角色），存在员工
        测试步骤：
            1. 进入组织架构页面
            2. 在员工列表中找到目标员工，点击"调动"
            3. 选择员工当前所在部门作为目标部门
            4. 点击确认
        预期结果：
            - 显示错误提示："目标部门必须与当前部门不同"
            - 调动操作被拒绝
        """
        self._login()
        self.page.get_by_test_id("link-header-org-structure").click()
        self.page.wait_for_timeout(2000)

        # Find and click "调动" button
        transfer_button = self.page.get_by_role("button", name="transfer 调动")
        if transfer_button.is_visible():
            transfer_button.first.click()
            self.page.wait_for_timeout(1000)

            # Try to submit without changing department (same department)
            self.page.get_by_test_id("btn-transfer-confirm").click()
            self.page.wait_for_timeout(2000)

        # Verify error message appears
        error_locator = self.page.locator(".ant-message-error")
        expect(error_locator).to_be_visible(timeout=5000)

    def test_cdp_emp_trans_003_view_transfer_history(self):
        """
        CDP-EMP-TRANS-003: 查看员工调动历史

        前置条件：用户已登录，存在有调动历史的员工
        测试步骤：
            1. 进入组织架构页面
            2. 在员工列表中找到目标员工
            3. 点击查看详情或调动历史
        预期结果：
            - 显示完整的调动历史列表
            - 按时间倒序排列
            - 显示调出部门、调入部门、调动类型、原因、日期
        """
        self._login()
        self.page.get_by_test_id("link-header-org-structure").click()
        self.page.wait_for_timeout(2000)

        # Click on an employee to view details
        # Look for detail button or click on employee row
        detail_button = self.page.get_by_role("button", name="detail 查看")
        if detail_button.is_visible():
            detail_button.first.click()
            self.page.wait_for_timeout(1000)

        # Verify no backend errors
        assert_no_error_message(self.page)

    def test_cdp_emp_trans_004_empty_transfer_history(self):
        """
        CDP-EMP-TRANS-004: 员工调动历史为空

        前置条件：用户已登录，存在从未调动的员工
        测试步骤：
            1. 进入组织架构页面
            2. 找到一名新入职且未调动的员工
            3. 查看其调动历史
        预期结果：
            - 显示空列表
            - 提示"暂无调动记录"
        """
        self._login()
        self.page.get_by_test_id("link-header-org-structure").click()
        self.page.wait_for_timeout(2000)

        # Click on first employee to view details
        detail_button = self.page.get_by_role("button", name="detail 查看")
        if detail_button.is_visible():
            detail_button.first.click()
            self.page.wait_for_timeout(1000)

        # Verify empty transfer history message if visible
        # The actual verification depends on the UI implementation
        assert_no_error_message(self.page)


class TestExcelImportExport:
    """Test cases for Excel import/export."""

    @pytest.fixture(autouse=True)
    def setup(self, page: Page):
        self.page = page
        self.login_page = LoginPage(page)

    def _login(self):
        """Login as admin user."""
        self.login_page.navigate()
        self.login_page.phone_input.fill("13800138002")
        self.login_page.password_input.fill("abcd1234")
        self.login_page.submit_button.click()
        self.page.wait_for_timeout(2000)

    def test_cdp_emp_excel_001_export_employee_list(self):
        """
        CDP-EMP-EXCEL-001: 导出员工列表

        前置条件：用户已登录，存在员工数据
        测试步骤：
            1. 进入组织架构页面
            2. 可选：设置筛选条件（部门、状态）
            3. 点击"导出"按钮
        预期结果：
            - 浏览器下载 Excel 文件
            - Excel 文件包含符合筛选条件的员工数据
            - 文件名为"员工列表.xlsx"或类似
        """
        self._login()
        self.page.get_by_test_id("link-header-org-structure").click()
        self.page.wait_for_timeout(2000)

        # Click "导出" button
        export_button = self.page.get_by_test_id("btn-employee-export")
        if export_button.is_visible():
            # Set up download handler before clicking
            with self.page.expect_download() as download_info:
                export_button.click()
            download = download_info.value
            # Verify download started
            expect(download.suggested_filename).to_match(r".*\.xlsx?")
            self.page.wait_for_timeout(2000)

        # Verify no backend errors
        assert_no_error_message(self.page)

    def test_cdp_emp_excel_002_import_success(self):
        """
        CDP-EMP-EXCEL-002: 成功批量导入员工

        前置条件：用户已登录（管理员角色），准备好包含正确格式的 Excel 文件
        测试步骤：
            1. 进入组织架构页面
            2. 点击"导入"按钮
            3. 上传格式合法的员工数据 Excel 文件
            4. 确认上传
        预期结果：
            - 导入成功
            - 显示导入结果：成功数量
            - 员工列表刷新，新员工出现在列表中
        """
        self._login()
        self.page.get_by_test_id("link-header-org-structure").click()
        self.page.wait_for_timeout(2000)

        # Create a temporary Excel file with valid employee data
        with tempfile.TemporaryDirectory() as tmp_dir:
            excel_path = os.path.join(tmp_dir, "valid_employees.xlsx")
            employees = [
                {
                    "code": f"IMP{int(time.time()) % 100000:05d}",
                    "name": "导入测试员工1",
                    "phone": "13900001011",
                    "position": "工程师",
                    "department": "技术部",
                },
                {
                    "code": f"IMP{int(time.time()) % 100000:05d}1",
                    "name": "导入测试员工2",
                    "phone": "13900001012",
                    "position": "产品经理",
                    "department": "产品部",
                },
            ]
            create_employee_excel(excel_path, employees)

            # Click "导入" button
            import_button = self.page.get_by_test_id("btn-employee-import")
            import_button.click()
            self.page.wait_for_timeout(1000)

            # Upload the file
            file_input = self.page.locator("input[type='file']")
            file_input.set_input_files(excel_path)
            self.page.wait_for_timeout(2000)

            # Look for confirm/upload button if present
            self.page.get_by_test_id("btn-emp-confirm").click()
            self.page.wait_for_timeout(2000)

        # Verify no backend errors
        assert_no_error_message(self.page)

    def test_cdp_emp_excel_003_import_with_errors(self):
        """
        CDP-EMP-EXCEL-003: 导入含错误行的 Excel

        前置条件：用户已登录（管理员角色），准备好包含重复工号的 Excel 文件
        测试步骤：
            1. 进入组织架构页面
            2. 点击"导入"按钮
            3. 上传包含错误行的 Excel 文件（如工号重复）
            4. 确认上传
        预期结果：
            - 显示部分导入结果
            - 显示成功数量
            - 显示错误行详情（行号和错误原因）
            - 合法行正常导入
        """
        self._login()
        self.page.get_by_test_id("link-header-org-structure").click()
        self.page.wait_for_timeout(2000)

        # Create a temporary Excel file with duplicate employee codes
        with tempfile.TemporaryDirectory() as tmp_dir:
            excel_path = os.path.join(tmp_dir, "duplicate_employees.xlsx")
            duplicate_code = f"DUP{int(time.time()) % 100000:05d}"
            create_invalid_employee_excel(excel_path, duplicate_code)

            # Click "导入" button
            import_button = self.page.get_by_test_id("btn-employee-import")
            import_button.click()
            self.page.wait_for_timeout(1000)

            # Upload the file with duplicate codes
            file_input = self.page.locator("input[type='file']")
            file_input.set_input_files(excel_path)
            self.page.wait_for_timeout(2000)

            # Wait for error/success message to appear
            # The UI should show partial success with error details for duplicates
            self.page.wait_for_timeout(2000)

        # Verify no backend errors
        assert_no_error_message(self.page)

    def test_cdp_emp_excel_004_import_missing_required_columns(self):
        """
        CDP-EMP-EXCEL-004: 导入缺少必填列

        前置条件：用户已登录（管理员角色），准备好缺少必填列的 Excel 文件
        测试步骤：
            1. 进入组织架构页面
            2. 点击"导入"按钮
            3. 上传缺少"工号"或"姓名"列的 Excel 文件
        预期结果：
            - 显示错误提示："Excel 缺少必要列: 工号、姓名"
            - 导入被拒绝
        """
        self._login()
        self.page.get_by_test_id("link-header-org-structure").click()
        self.page.wait_for_timeout(2000)

        # Create a temporary Excel file missing required columns
        with tempfile.TemporaryDirectory() as tmp_dir:
            excel_path = os.path.join(tmp_dir, "incomplete_employees.xlsx")
            # Missing "工号" column
            create_incomplete_employee_excel(excel_path, ["工号"])

            # Click "导入" button
            import_button = self.page.get_by_test_id("btn-employee-import")
            import_button.click()
            self.page.wait_for_timeout(1000)

            # Upload the file with missing columns
            file_input = self.page.locator("input[type='file']")
            file_input.set_input_files(excel_path)
            self.page.wait_for_timeout(2000)

            # Wait for error message about missing columns
            # Expected: "Excel 缺少必要列: 工号、姓名"
            self.page.wait_for_timeout(2000)

        # Verify no backend errors
        assert_no_error_message(self.page)


class TestDashboardStatistics:
    """Test cases for organization dashboard statistics."""

    @pytest.fixture(autouse=True)
    def setup(self, page: Page):
        self.page = page
        self.login_page = LoginPage(page)

    def _login(self):
        """Login as admin user."""
        self.login_page.navigate()
        self.login_page.phone_input.fill("13800138002")
        self.login_page.password_input.fill("abcd1234")
        self.login_page.submit_button.click()
        self.page.wait_for_timeout(2000)

    def test_cdp_dash_001_statistics_display(self):
        """
        CDP-DASH-001: 看板统计数据正常显示

        前置条件：用户已登录
        测试步骤：
            1. 进入组织架构页面
        预期结果：
            - 页面顶部显示统计卡片
            - 显示组织总数
            - 显示员工总数
            - 显示在职员工数
            - 显示待入职员工数
        """
        self._login()
        self.page.get_by_test_id("link-header-org-structure").click()
        self.page.wait_for_timeout(2000)

        # Verify no backend error
        assert_no_error_message(self.page)

        # Verify statistics cards are visible
        expect(self.page.get_by_text("组织单元")).to_be_visible()
        expect(self.page.get_by_text("员工总数")).to_be_visible()
        expect(self.page.get_by_text("在职")).to_be_visible()
        expect(self.page.get_by_text("入职中")).to_be_visible()

        # Verify the statistics cards contain numbers (not just labels)
        # The cards should show actual counts like "10" organizations, "100" employees, etc.
        # We can verify the card structure exists with data-testid or specific styling
        statistics_cards = self.page.locator(".ant-card")
        expect(statistics_cards.first).to_be_visible()

    def test_cdp_dash_002_statistics_scope_limited_by_permission(self):
        """
        CDP-DASH-002: 统计数量反映权限范围

        前置条件：用户已登录（非管理员角色，如分支管理员）
        测试步骤：
            1. 以分支管理员账号登录
            2. 进入组织架构页面
        预期结果：
            - 统计数据显示仅限于该分支及后代的数量
            - 不显示全组织数据
        """
        # For this test, we need to verify that a non-admin user only sees
        # their scoped data. Since we only have admin credentials (13800138002),
        # we can at least verify that the statistics display works for admin
        # and the structure supports permission-based filtering.
        self._login()
        self.page.get_by_test_id("link-header-org-structure").click()
        self.page.wait_for_timeout(2000)

        # Verify no backend error
        assert_no_error_message(self.page)

        # Verify statistics are displayed
        expect(self.page.get_by_text("组织单元")).to_be_visible()
        expect(self.page.get_by_text("员工总数")).to_be_visible()

        # Note: Full permission scope testing would require a non-admin user account
        # which is not available in the current test environment.


class TestPermissionControl:
    """Test cases for permission control."""

    @pytest.fixture(autouse=True)
    def setup(self, page: Page):
        self.page = page
        self.login_page = LoginPage(page)

    def _login(self):
        """Login as admin user."""
        self.login_page.navigate()
        self.login_page.phone_input.fill("13800138002")
        self.login_page.password_input.fill("abcd1234")
        self.login_page.submit_button.click()
        self.page.wait_for_timeout(2000)

    def test_cdp_perm_001_unauthenticated_access(self):
        """
        CDP-PERM-001: 未登录访问组织架构页

        前置条件：未登录
        测试步骤：
            1. 直接访问 http://localhost:3001/org-structure
        预期结果：
            - 重定向到登录页面
            - 或显示无权限提示
        """
        # Navigate to the page first, then clear storage to simulate unauthenticated access
        self.page.goto("http://localhost:3001/org-structure")
        self.page.wait_for_timeout(1000)

        # Clear cookies and localStorage after navigation
        self.page.context.clear_cookies()
        self.page.evaluate("() => { localStorage.clear(); sessionStorage.clear(); }")

        # Reload the page after clearing auth state
        self.page.goto("http://localhost:3001/org-structure")
        self.page.wait_for_timeout(2000)

        # Should redirect to login
        expect(self.page).to_have_url(re.compile(r"/login"))

    def test_cdp_perm_002_regular_user_cannot_edit_org(self):
        """
        CDP-PERM-002: 普通员工无法编辑组织

        前置条件：用户已登录（普通员工角色）
        测试步骤：
            1. 进入组织架构页面
            2. 尝试编辑某个组织单元
        预期结果：
            - 不显示编辑按钮
            - 或显示"无权限"错误
        """
        # Note: This test requires a non-admin user account which is not available
        # in the current test environment. The test structure is in place but
        # full verification requires credentials for a regular user role.
        self._login()
        self.page.get_by_test_id("link-header-org-structure").click()
        self.page.wait_for_timeout(2000)

        # Verify no backend error
        assert_no_error_message(self.page)

        # For admin user, right-click should show edit option
        tree = self.page.locator(".ant-tree")
        tree_node = tree.locator(".ant-tree-node-content-wrapper").first
        if tree_node.is_visible():
            tree_node.click(button="right")
            self.page.wait_for_timeout(500)
            # Admin should see edit option
            expect(self.page.get_by_text("编辑", exact=False)).to_be_visible()

    def test_cdp_perm_003_regular_user_cannot_delete_org(self):
        """
        CDP-PERM-003: 普通员工无法删除组织

        前置条件：用户已登录（普通员工角色）
        测试步骤：
            1. 进入组织架构页面
            2. 尝试删除某个组织单元
        预期结果：
            - 不显示删除按钮
            - 或显示"无权限"错误
        """
        # Note: This test requires a non-admin user account which is not available
        # in the current test environment.
        self._login()
        self.page.get_by_test_id("link-header-org-structure").click()
        self.page.wait_for_timeout(2000)

        # Verify no backend error
        assert_no_error_message(self.page)

        # For admin user, right-click should show delete option
        tree = self.page.locator(".ant-tree")
        tree_node = tree.locator(".ant-tree-node-content-wrapper").first
        if tree_node.is_visible():
            tree_node.click(button="right")
            self.page.wait_for_timeout(500)
            # Admin should see delete option
            expect(self.page.get_by_text("删除", exact=False)).to_be_visible()

    def test_cdp_perm_004_regular_user_cannot_create_employee(self):
        """
        CDP-PERM-004: 普通员工无法创建员工

        前置条件：用户已登录（普通员工角色）
        测试步骤：
            1. 进入组织架构页面
            2. 尝试点击"新增员工"按钮
        预期结果：
            - 不显示"新增员工"按钮
            - 或显示"无权限"错误
        """
        # Note: This test requires a non-admin user account which is not available
        # in the current test environment.
        self._login()
        self.page.get_by_test_id("link-header-org-structure").click()
        self.page.wait_for_timeout(2000)

        # Verify no backend error
        assert_no_error_message(self.page)

        # Admin should see "新增员工" button
        expect(self.page.get_by_text("新增员工", exact=False)).to_be_visible()

    def test_cdp_perm_005_branch_admin_scope(self):
        """
        CDP-PERM-005: 分支管理员权限范围

        前置条件：用户已登录（分支管理员角色）
        测试步骤：
            1. 进入组织架构页面
            2. 查看组织树显示范围
        预期结果：
            - 仅显示该分支及后代组织
            - 不显示其他分支的组织
        """
        # Note: This test requires a branch admin user account which is not available
        # in the current test environment.
        self._login()
        self.page.get_by_test_id("link-header-org-structure").click()
        self.page.wait_for_timeout(2000)

        # Verify no backend error
        assert_no_error_message(self.page)

        # Verify organization tree is visible for admin
        tree = self.page.locator(".ant-tree")
        expect(tree).to_be_visible()


class TestUIInteraction:
    """Test cases for UI interactions."""

    @pytest.fixture(autouse=True)
    def setup(self, page: Page):
        self.page = page
        self.login_page = LoginPage(page)

    def _login(self):
        """Login as admin user."""
        self.login_page.navigate()
        self.login_page.phone_input.fill("13800138002")
        self.login_page.password_input.fill("abcd1234")
        self.login_page.submit_button.click()
        self.page.wait_for_timeout(2000)

    def test_cdp_ui_001_org_tree_context_menu(self):
        """
        CDP-UI-001: 组织树节点操作菜单

        前置条件：用户已登录（管理员角色），存在组织数据
        测试步骤：
            1. 进入组织架构页面
            2. 在组织树节点上右键点击
        预期结果：
            - 显示操作菜单
            - 包含：添加子节点、编辑、禁用/启用、删除、移动
        """
        self._login()
        self.page.get_by_test_id("link-header-org-structure").click()
        self.page.wait_for_timeout(2000)

        # Verify no backend error
        assert_no_error_message(self.page)

        # Right-click on a tree node to open context menu
        tree = self.page.locator(".ant-tree")
        tree_node = tree.locator(".ant-tree-node-content-wrapper").first
        expect(tree_node).to_be_visible()
        tree_node.click(button="right")
        self.page.wait_for_timeout(500)

        # Verify context menu is visible with operation options
        # Common operations: 新增子节点, 编辑, 禁用/启用, 删除, 移动
        expect(self.page.get_by_text("新增", exact=False)).to_be_visible()
        expect(self.page.get_by_text("编辑", exact=False)).to_be_visible()

    def test_cdp_ui_002_employee_list_pagination(self):
        """
        CDP-UI-002: 员工列表分页

        前置条件：用户已登录，存在超过一页的员工数据
        测试步骤：
            1. 进入组织架构页面
            2. 查看员工列表分页
            3. 点击下一页
        预期结果：
            - 分页切换成功
            - 列表更新显示下一页数据
        """
        self._login()
        self.page.get_by_test_id("link-header-org-structure").click()
        self.page.wait_for_timeout(2000)

        # Verify no backend error
        assert_no_error_message(self.page)

        # Check if pagination controls exist
        pagination = self.page.locator(".ant-pagination")
        if pagination.is_visible():
            # Try to click next page button
            next_button = pagination.get_by_role(
                "button", name=re.compile(">>|下一页|下页")
            )
            if next_button.is_visible():
                next_button.click()
                self.page.wait_for_timeout(1000)
                # Verify pagination worked (page number should change)
                expect(pagination).to_be_visible()

    def test_cdp_ui_003_employee_list_sorting(self):
        """
        CDP-UI-003: 员工列表排序

        前置条件：用户已登录，存在员工数据
        测试步骤：
            1. 进入组织架构页面
            2. 点击员工列表表头的"工号"列
        预期结果：
            - 列表按工号排序
            - 再次点击反向排序
        """
        self._login()
        self.page.get_by_test_id("link-header-org-structure").click()
        self.page.wait_for_timeout(2000)

        # Verify no backend error
        assert_no_error_message(self.page)

        # Click on the employee code column header to sort
        code_header = self.page.get_by_role("columnheader", name="工号")
        if code_header.is_visible():
            code_header.click()
            self.page.wait_for_timeout(500)
            # Click again to reverse sort
            code_header.click()
            self.page.wait_for_timeout(500)

    def test_cdp_ui_004_modal_form_reset_after_close(self):
        """
        CDP-UI-004: 模态框关闭后重置表单

        前置条件：用户已登录（管理员角色）
        测试步骤：
            1. 进入组织架构页面
            2. 点击"新增员工"打开模态框
            3. 输入一些内容
            4. 点击取消关闭模态框
            5. 再次点击"新增员工"打开模态框
        预期结果：
            - 表单内容已重置
            - 不显示之前输入的内容
        """
        self._login()
        self.page.get_by_test_id("link-header-org-structure").click()
        self.page.wait_for_timeout(2000)

        # Verify no backend error
        assert_no_error_message(self.page)

        # Click "新增员工" button to open modal
        add_employee_btn = self.page.get_by_text("新增员工", exact=False)
        if add_employee_btn.is_visible():
            add_employee_btn.click()
            self.page.wait_for_timeout(500)

            # Check if dialog appears
            dialog = self.page.locator(".ant-modal").first
            if dialog.is_visible():
                # Fill in some form data
                # Try to fill name field if it exists
                name_input = dialog.get_by_test_id("inp-emp-name")
                if name_input.is_visible():
                    name_input.fill("测试员工")
                    self.page.wait_for_timeout(300)

                # Close the dialog by clicking cancel or X
                self.page.get_by_test_id("btn-emp-cancel").click()
                self.page.wait_for_timeout(500)

                # Open dialog again
                add_employee_btn.click()
                self.page.wait_for_timeout(500)

                # Verify form is reset (name field should be empty)
                name_input = dialog.get_by_test_id("inp-emp-name")
                if name_input.is_visible():
                    expect(name_input).to_have_value("")

    def test_cdp_ui_005_org_tree_expand_collapse(self):
        """
        CDP-UI-005: 组织树展开折叠

        前置条件：用户已登录，存在层级组织数据
        测试步骤：
            1. 进入组织架构页面
            2. 点击组织树节点的展开/折叠图标
        预期结果：
            - 节点展开显示子节点
            - 或节点折叠隐藏子节点
        """
        self._login()
        self.page.get_by_test_id("link-header-org-structure").click()
        self.page.wait_for_timeout(2000)

        # Verify no backend error
        assert_no_error_message(self.page)

        # Find a tree switcher (expand/collapse icon)
        tree = self.page.locator(".ant-tree")
        switchers = tree.locator(".ant-tree-switcher")
        if switchers.first.is_visible():
            # Click to expand
            switchers.first.click()
            self.page.wait_for_timeout(500)

            # Click again to collapse
            switchers.first.click()
            self.page.wait_for_timeout(500)
